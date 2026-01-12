"""Excel report generator - infrastructure adapter for xlsx export."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from swen.application.dtos.analytics.analytics_dto import MonthComparisonResult
from swen.application.dtos.export_dto import AccountExportDTO
from swen.application.dtos.export_report_dto import (
    DashboardSummaryDTO,
    ExportReportData,
    MappingExportRowDTO,
    TransactionExportRowDTO,
)


class ExcelStyles:
    """Centralized style definitions for Excel formatting."""

    # Colors
    PRIMARY_BLUE = "1E3A5F"  # Dark blue
    ACCENT_TEAL = "0D9488"  # Teal accent
    SUCCESS_GREEN = "059669"  # Green for positive
    WARNING_AMBER = "D97706"  # Amber for negative
    HEADER_BG = "1E3A5F"
    HEADER_FG = "FFFFFF"
    SUBHEADER_BG = "E5E7EB"
    ALT_ROW_BG = "F9FAFB"
    BORDER_COLOR = "D1D5DB"

    # Fonts
    TITLE_FONT = Font(name="Calibri", size=24, bold=True, color=PRIMARY_BLUE)
    SUBTITLE_FONT = Font(name="Calibri", size=12, color="6B7280")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color=PRIMARY_BLUE)
    METRIC_LABEL_FONT = Font(name="Calibri", size=10, color="6B7280")
    METRIC_VALUE_FONT = Font(name="Calibri", size=14, bold=True, color=PRIMARY_BLUE)
    LARGE_METRIC_FONT = Font(name="Calibri", size=18, bold=True, color=PRIMARY_BLUE)
    BODY_FONT = Font(name="Calibri", size=10)
    POSITIVE_FONT = Font(name="Calibri", size=10, color=SUCCESS_GREEN)
    NEGATIVE_FONT = Font(name="Calibri", size=10, color=WARNING_AMBER)

    # Fills
    HEADER_FILL = PatternFill(
        start_color=HEADER_BG,
        end_color=HEADER_BG,
        fill_type="solid",
    )
    SUBHEADER_FILL = PatternFill(
        start_color=SUBHEADER_BG,
        end_color=SUBHEADER_BG,
        fill_type="solid",
    )
    ALT_ROW_FILL = PatternFill(
        start_color=ALT_ROW_BG,
        end_color=ALT_ROW_BG,
        fill_type="solid",
    )

    # Borders
    THIN_BORDER = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    # Alignments
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")
    WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


class ExcelReportGenerator:
    """Generates beautifully formatted Excel reports from ExportReportData.

    Creates a multi-sheet workbook with:
    1. Dashboard - Summary metrics and overview
    2. Transactions - Full transaction data with metadata
    3. Accounts - Chart of accounts
    4. Mappings - Bank account mappings
    """

    def __init__(self):
        self._styles = ExcelStyles()

    def generate(self, data: ExportReportData) -> bytes:
        wb = Workbook()

        # Remove default sheet and create our sheets
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        # Sheet 1: Dashboard (overview)
        self._create_dashboard_sheet(wb, data.summary)

        # Sheet 2: Transactions
        self._create_transactions_sheet(wb, data.transactions)

        # Sheet 3: Accounts
        self._create_accounts_sheet(wb, data.accounts)

        # Sheet 4: Bank Mappings
        self._create_mappings_sheet(wb, data.mappings)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_dashboard_sheet(
        self,
        wb: Workbook,
        summary: DashboardSummaryDTO,
    ) -> None:
        ws = wb.create_sheet(title="Dashboard")
        row = 1

        row = self._add_dashboard_header(ws, row, summary)
        row = self._add_summary_metrics(ws, row, summary)

        if summary.month_comparison:
            row = self._add_month_comparison(ws, row, summary.month_comparison)

        if summary.account_balances:
            row = self._add_account_balances(ws, row, summary.account_balances)

        if summary.top_expenses:
            row = self._add_top_expenses(ws, row, summary.top_expenses)

        if summary.net_worth_trend:
            self._add_net_worth_trend(ws, row, summary.net_worth_trend)

        self._set_dashboard_column_widths(ws)

    def _add_dashboard_header(
        self,
        ws: Worksheet,
        row: int,
        summary: DashboardSummaryDTO,
    ) -> int:
        """Add title and metadata header to dashboard."""
        ws.merge_cells(f"A{row}:F{row}")
        title_cell = ws.cell(row=row, column=1, value=f"🦜 {summary.report_title}")
        title_cell.font = self._styles.TITLE_FONT
        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row=row, column=1, value=f"Period: {summary.period_label}")
        ws.cell(row=row, column=1).font = self._styles.SUBTITLE_FONT
        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        generated = summary.generated_at.strftime("%d %B %Y, %H:%M")
        ws.cell(row=row, column=1, value=f"Generated: {generated}")
        ws.cell(row=row, column=1).font = self._styles.SUBTITLE_FONT
        return row + 2

    def _add_summary_metrics(
        self,
        ws: Worksheet,
        row: int,
        summary: DashboardSummaryDTO,
    ) -> int:
        """Add primary and secondary metrics grid."""
        row = self._add_section_header(ws, row, "SUMMARY")
        row += 1

        # Primary metrics row
        metrics = [
            ("Total Income", summary.total_income),
            ("Total Expenses", summary.total_expenses),
            ("Net Income", summary.net_income),
        ]

        for col, (label, value) in enumerate(metrics, start=1):
            ws.cell(row=row, column=col, value=label)
            ws.cell(row=row, column=col).font = self._styles.METRIC_LABEL_FONT
            ws.cell(row=row, column=col).alignment = self._styles.CENTER

            value_cell = ws.cell(row=row + 1, column=col)
            value_cell.value = float(value)  # type: ignore[union-attr]
            value_cell.number_format = "#,##0.00 €"
            value_cell.font = (
                self._styles.LARGE_METRIC_FONT
                if col == 3
                else self._styles.METRIC_VALUE_FONT
            )
            value_cell.alignment = self._styles.CENTER

        row += 3

        # Secondary metrics row
        secondary_metrics = [
            ("Savings Rate", float(summary.savings_rate) / 100, "0.0%"),
            ("Net Worth", float(summary.net_worth), "#,##0.00 €"),
            ("Transactions", summary.transaction_count, None),
        ]

        for col, (label, value, fmt) in enumerate(secondary_metrics, start=1):
            ws.cell(row=row, column=col, value=label)
            ws.cell(row=row, column=col).font = self._styles.METRIC_LABEL_FONT
            value_cell = ws.cell(row=row + 1, column=col)
            value_cell.value = value  # type: ignore[union-attr]
            if fmt:
                value_cell.number_format = fmt
            value_cell.font = self._styles.METRIC_VALUE_FONT
            value_cell.alignment = self._styles.CENTER

        return row + 4

    def _add_month_comparison(
        self,
        ws: Worksheet,
        row: int,
        comp: MonthComparisonResult,
    ) -> int:
        """Add month-over-month comparison table."""
        row = self._add_section_header(ws, row, "MONTH-OVER-MONTH COMPARISON")
        row += 1

        headers = ["Metric", comp.previous_month, comp.current_month, "Change"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self._styles.SUBHEADER_FONT
            cell.fill = self._styles.SUBHEADER_FILL
        row += 1

        comparison_rows = [
            (
                "Income",
                comp.previous_income,
                comp.current_income,
                comp.income_change_percentage,
            ),
            (
                "Spending",
                comp.previous_spending,
                comp.current_spending,
                comp.spending_change_percentage,
            ),
            (
                "Net Income",
                comp.previous_net,
                comp.current_net,
                comp.net_change_percentage,
            ),
        ]

        for label, prev, curr, change in comparison_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=float(prev)).number_format = "#,##0.00 €"
            ws.cell(row=row, column=3, value=float(curr)).number_format = "#,##0.00 €"
            change_cell = ws.cell(row=row, column=4)
            change_cell.value = float(change) / 100  # type: ignore[union-attr]
            change_cell.number_format = "+0.0%;-0.0%"
            change_cell.font = (
                self._styles.POSITIVE_FONT
                if change >= 0
                else self._styles.NEGATIVE_FONT
            )
            row += 1

        return row + 1

    def _add_account_balances(
        self,
        ws: Worksheet,
        row: int,
        balances: list,
    ) -> int:
        """Add account balances section."""
        row = self._add_section_header(ws, row, "ACCOUNT BALANCES")
        row += 1

        for balance in balances[:8]:  # Top 8 accounts
            ws.cell(row=row, column=1, value=balance.account_name)
            ws.cell(row=row, column=1).font = self._styles.BODY_FONT
            value_cell = ws.cell(row=row, column=2)
            value_cell.value = float(balance.balance)  # type: ignore[union-attr]
            value_cell.number_format = "#,##0.00 €"
            value_cell.font = self._styles.BODY_FONT
            value_cell.alignment = self._styles.RIGHT
            row += 1

        return row + 1

    def _add_top_expenses(
        self,
        ws: Worksheet,
        row: int,
        expenses: list,
    ) -> int:
        """Add top expense categories table."""
        row = self._add_section_header(ws, row, "TOP EXPENSE CATEGORIES")
        row += 1

        headers = ["Category", "Amount", "% of Total"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self._styles.SUBHEADER_FONT
            cell.fill = self._styles.SUBHEADER_FILL
        row += 1

        for expense in expenses[:10]:
            ws.cell(row=row, column=1, value=expense.category)
            ws.cell(
                row=row,
                column=2,
                value=float(expense.amount),
            ).number_format = "#,##0.00 €"
            ws.cell(
                row=row,
                column=3,
                value=float(expense.percentage) / 100,
            ).number_format = "0.0%"
            row += 1

        return row + 1

    def _add_net_worth_trend(
        self,
        ws: Worksheet,
        row: int,
        trend: list,
    ) -> int:
        """Add net worth trend section."""
        row = self._add_section_header(ws, row, "NET WORTH TREND")
        row += 1

        for dp in trend:
            ws.cell(row=row, column=1, value=dp.period_label)
            ws.cell(
                row=row,
                column=2,
                value=float(dp.value),
            ).number_format = "#,##0.00 €"
            row += 1

        return row

    def _set_dashboard_column_widths(self, ws: Worksheet) -> None:
        """Set column widths for dashboard sheet."""
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

    def _add_section_header(self, ws: Worksheet, row: int, title: str) -> int:
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self._styles.SUBHEADER_FONT
        cell.fill = self._styles.SUBHEADER_FILL
        return row + 1

    def _create_transactions_sheet(
        self,
        wb: Workbook,
        transactions: list[TransactionExportRowDTO],
    ) -> None:
        ws = wb.create_sheet(title="Transactions")

        # Headers
        headers = TransactionExportRowDTO.column_headers()
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._styles.HEADER_FONT
            cell.fill = self._styles.HEADER_FILL
            cell.alignment = self._styles.CENTER

        # Data rows
        for row_idx, txn in enumerate(transactions, start=2):
            row_data = txn.to_row()
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = self._styles.BODY_FONT

                # Apply currency format to amount column
                if col == 4:  # Amount
                    cell.number_format = "#,##0.00"
                    cell.alignment = self._styles.RIGHT

                # Alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = self._styles.ALT_ROW_FILL

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-fit columns (approximate widths)
        column_widths = [
            12,
            40,
            25,
            12,
            8,
            25,
            25,
            10,
            12,
            25,
            25,
            50,
            20,
            12,
            20,
            12,
            10,
        ]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = min(width, 50)

    def _create_accounts_sheet(
        self,
        wb: Workbook,
        accounts: list[AccountExportDTO],
    ) -> None:
        ws = wb.create_sheet(title="Accounts")

        headers = ["Account Number", "Name", "Type", "Currency", "Active", "Parent ID"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._styles.HEADER_FONT
            cell.fill = self._styles.HEADER_FILL
            cell.alignment = self._styles.CENTER

        # Sort accounts by account number
        sorted_accounts = sorted(accounts, key=lambda a: a.account_number)

        for row_idx, acc in enumerate(sorted_accounts, start=2):
            ws.cell(row=row_idx, column=1, value=acc.account_number)
            ws.cell(row=row_idx, column=2, value=acc.name)
            ws.cell(row=row_idx, column=3, value=acc.type.title())
            ws.cell(row=row_idx, column=4, value=acc.currency)
            ws.cell(row=row_idx, column=5, value="Yes" if acc.is_active else "No")
            ws.cell(row=row_idx, column=6, value=acc.parent_id or "")

            for col in range(1, 7):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = self._styles.BODY_FONT
                if row_idx % 2 == 0:
                    cell.fill = self._styles.ALT_ROW_FILL

        # Freeze header
        ws.freeze_panes = "A2"

        # Column widths
        widths = [15, 30, 12, 10, 10, 40]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_mappings_sheet(
        self,
        wb: Workbook,
        mappings: list[MappingExportRowDTO],
    ) -> None:
        ws = wb.create_sheet(title="Bank Mappings")

        # Headers from DTO
        headers = MappingExportRowDTO.column_headers()
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._styles.HEADER_FONT
            cell.fill = self._styles.HEADER_FILL
            cell.alignment = self._styles.CENTER

        # Data rows
        for row_idx, mapping in enumerate(mappings, start=2):
            row_data = mapping.to_row()
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = self._styles.BODY_FONT
                if row_idx % 2 == 0:
                    cell.fill = self._styles.ALT_ROW_FILL

        # Freeze header
        ws.freeze_panes = "A2"

        # Column widths
        widths = [28, 35, 30, 22]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
