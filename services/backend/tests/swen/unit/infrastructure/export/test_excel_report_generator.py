"""Unit tests for ExcelReportGenerator.

These tests verify the Excel report generation without requiring
a database or other infrastructure.
"""

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from swen.application.dtos.analytics import (
    BreakdownItem,
    MonthComparisonResult,
    TimeSeriesDataPoint,
)
from swen.application.dtos.export_dto import AccountExportDTO
from swen.application.dtos.export_report_dto import (
    AccountBalanceSummary,
    DashboardSummaryDTO,
    ExportReportData,
    MappingExportRowDTO,
    TransactionExportRowDTO,
)
from swen.infrastructure.export import ExcelReportGenerator


class TestExcelReportGenerator:
    """Test suite for ExcelReportGenerator."""

    @pytest.fixture
    def generator(self) -> ExcelReportGenerator:
        """Create a fresh generator instance."""
        return ExcelReportGenerator()

    @pytest.fixture
    def sample_summary(self) -> DashboardSummaryDTO:
        """Create sample dashboard summary data."""
        return DashboardSummaryDTO(
            report_title="SWEN Financial Report",
            period_label="December 2024",
            start_date=date(2024, 12, 1),
            end_date=date(2024, 12, 31),
            generated_at=datetime(2024, 12, 18, 14, 30, 0, tzinfo=timezone.utc),
            currency="EUR",
            total_income=Decimal("5000.00"),
            total_expenses=Decimal("3500.00"),
            net_income=Decimal("1500.00"),
            savings_rate=Decimal("30.0"),
            net_worth=Decimal("25000.00"),
            account_balances=[
                AccountBalanceSummary(
                    account_name="DKB Checking",
                    account_number="1000",
                    balance=Decimal("5000.00"),
                    currency="EUR",
                ),
                AccountBalanceSummary(
                    account_name="Savings",
                    account_number="1100",
                    balance=Decimal("20000.00"),
                    currency="EUR",
                ),
            ],
            top_expenses=[
                BreakdownItem(
                    category="Rent",
                    amount=Decimal("1200.00"),
                    percentage=Decimal("34.3"),
                    account_id="uuid-1",
                ),
                BreakdownItem(
                    category="Groceries",
                    amount=Decimal("500.00"),
                    percentage=Decimal("14.3"),
                    account_id="uuid-2",
                ),
            ],
            month_comparison=MonthComparisonResult(
                current_month="December 2024",
                previous_month="November 2024",
                currency="EUR",
                current_income=Decimal("5000.00"),
                previous_income=Decimal("4800.00"),
                income_change=Decimal("200.00"),
                income_change_percentage=Decimal("4.2"),
                current_spending=Decimal("3500.00"),
                previous_spending=Decimal("3800.00"),
                spending_change=Decimal("-300.00"),
                spending_change_percentage=Decimal("-7.9"),
                current_net=Decimal("1500.00"),
                previous_net=Decimal("1000.00"),
                net_change=Decimal("500.00"),
                net_change_percentage=Decimal("50.0"),
                category_comparisons=[],
            ),
            net_worth_trend=[
                TimeSeriesDataPoint(
                    period="2024-11",
                    period_label="November 2024",
                    value=Decimal("24000.00"),
                ),
                TimeSeriesDataPoint(
                    period="2024-12",
                    period_label="December 2024",
                    value=Decimal("25000.00"),
                ),
            ],
            transaction_count=150,
            posted_count=145,
            draft_count=5,
        )

    @pytest.fixture
    def sample_transactions(self) -> list[TransactionExportRowDTO]:
        """Create sample transaction data."""
        return [
            TransactionExportRowDTO(
                id="txn-1",
                date="2024-12-15",
                description="REWE Supermarket",
                counterparty="REWE",
                amount=45.99,
                currency="EUR",
                debit_account="4200 - Groceries",
                credit_account="1000 - DKB Checking",
                status="posted",
                source="bank_import",
                source_iban="DE89370400440532013000",
                counterparty_iban="",
                is_internal_transfer=False,
                original_purpose="REWE SAGT DANKE",
                bank_reference="2024121512345",
                ai_suggested_account="Groceries",
                ai_confidence=0.95,
                ai_accepted=True,
            ),
            TransactionExportRowDTO(
                id="txn-2",
                date="2024-12-10",
                description="Transfer to Savings",
                counterparty="Savings",
                amount=500.00,
                currency="EUR",
                debit_account="1100 - Savings",
                credit_account="1000 - DKB Checking",
                status="posted",
                source="bank_import",
                source_iban="DE89370400440532013000",
                counterparty_iban="DE89370400440532013001",
                is_internal_transfer=True,
                original_purpose="UMBUCHUNG",
                bank_reference="2024121098765",
                ai_suggested_account="",
                ai_confidence=None,
                ai_accepted=None,
            ),
        ]

    @pytest.fixture
    def sample_accounts(self) -> list[AccountExportDTO]:
        """Create sample account data."""
        return [
            AccountExportDTO(
                id="acc-1",
                account_number="1000",
                name="DKB Checking",
                type="asset",
                currency="EUR",
                is_active=True,
                parent_id="",
                created_at="2024-01-01T00:00:00+00:00",
            ),
            AccountExportDTO(
                id="acc-2",
                account_number="4200",
                name="Groceries",
                type="expense",
                currency="EUR",
                is_active=True,
                parent_id="",
                created_at="2024-01-01T00:00:00+00:00",
            ),
        ]

    @pytest.fixture
    def sample_mappings(self) -> list[MappingExportRowDTO]:
        """Create sample mapping data."""
        return [
            MappingExportRowDTO(
                iban="DE89370400440532013000",
                bank_account_name="Malte Winckler - Girokonto",
                accounting_account_name="DKB Checking",
                accounting_account_number="1000",
                created_at="2024-01-01T00:00:00+00:00",
            ),
        ]

    @pytest.fixture
    def sample_report_data(
        self,
        sample_summary,
        sample_transactions,
        sample_accounts,
        sample_mappings,
    ) -> ExportReportData:
        """Create complete sample report data."""
        return ExportReportData(
            summary=sample_summary,
            transactions=sample_transactions,
            accounts=sample_accounts,
            mappings=sample_mappings,
        )

    def test_generate_returns_bytes(self, generator, sample_report_data):
        """Test that generate() returns bytes that can be read as Excel."""
        result = generator.generate(sample_report_data)

        assert isinstance(result, bytes)
        assert len(result) > 0

        # Verify it's valid Excel by loading it
        wb = load_workbook(BytesIO(result))
        assert wb is not None

    def test_generate_creates_four_sheets(self, generator, sample_report_data):
        """Test that the generated workbook has four sheets."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))

        assert len(wb.sheetnames) == 4
        assert "Dashboard" in wb.sheetnames
        assert "Transactions" in wb.sheetnames
        assert "Accounts" in wb.sheetnames
        assert "Bank Mappings" in wb.sheetnames

    def test_dashboard_sheet_contains_summary_metrics(
        self,
        generator,
        sample_report_data,
    ):
        """Test that Dashboard sheet contains the summary metrics."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Dashboard"]

        # Check title is present
        title_found = False
        for row in ws.iter_rows(max_row=5, max_col=6, values_only=True):
            for cell in row:
                if cell and "SWEN Financial Report" in str(cell):
                    title_found = True
                    break

        assert title_found, "Report title not found in Dashboard"

    def test_transactions_sheet_has_correct_headers(
        self,
        generator,
        sample_report_data,
    ):
        """Test that Transactions sheet has the correct headers."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Transactions"]

        headers = [cell.value for cell in ws[1]]
        expected_headers = TransactionExportRowDTO.column_headers()

        assert headers == expected_headers

    def test_transactions_sheet_contains_data(self, generator, sample_report_data):
        """Test that Transactions sheet contains the transaction data."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Transactions"]

        # Should have header + 2 data rows
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 3  # Header + 2 transactions

        # Check first transaction data
        data_row = rows[1]
        assert data_row[0] == "2024-12-15"  # Date
        assert data_row[1] == "REWE Supermarket"  # Description
        assert data_row[2] == "REWE"  # Counterparty

    def test_accounts_sheet_has_correct_headers(self, generator, sample_report_data):
        """Test that Accounts sheet has the correct headers."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Accounts"]

        headers = [cell.value for cell in ws[1]]
        expected = ["Account Number", "Name", "Type", "Currency", "Active", "Parent ID"]

        assert headers == expected

    def test_accounts_sheet_sorted_by_account_number(
        self,
        generator,
        sample_report_data,
    ):
        """Test that Accounts sheet is sorted by account number."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Accounts"]

        account_numbers = [
            str(row[0])
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True)
            if row[0]
        ]

        assert account_numbers == sorted(account_numbers)

    def test_mappings_sheet_has_correct_headers(self, generator, sample_report_data):
        """Test that Bank Mappings sheet has the correct headers."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Bank Mappings"]

        headers = [cell.value for cell in ws[1]]
        expected = MappingExportRowDTO.column_headers()

        assert headers == expected

    def test_mappings_sheet_shows_friendly_account_name(
        self,
        generator,
        sample_report_data,
    ):
        """Test that Bank Mappings shows friendly account names, not UUIDs."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Bank Mappings"]

        # Get first data row
        data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

        # Column 3 should be "1000 - DKB Checking", not a UUID
        linked_account = data_row[2]
        assert linked_account == "1000 - DKB Checking"
        assert isinstance(linked_account, str)
        assert "-" in linked_account  # Format: "number - name"

    def test_generate_with_empty_data(self, generator):
        """Test that generator handles empty data gracefully."""
        empty_summary = DashboardSummaryDTO(
            report_title="Empty Report",
            period_label="No Data",
            start_date=None,
            end_date=None,
            generated_at=datetime.now(tz=timezone.utc),
            currency="EUR",
        )

        empty_data = ExportReportData(
            summary=empty_summary,
            transactions=[],
            accounts=[],
            mappings=[],
        )

        result = generator.generate(empty_data)

        assert isinstance(result, bytes)
        wb = load_workbook(BytesIO(result))
        assert len(wb.sheetnames) == 4

    def test_header_rows_are_frozen(self, generator, sample_report_data):
        """Test that header rows are frozen for scrolling."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))

        # Transactions sheet should have frozen pane at A2
        transactions = wb["Transactions"]
        assert transactions.freeze_panes == "A2"

        # Accounts sheet should have frozen pane at A2
        accounts = wb["Accounts"]
        assert accounts.freeze_panes == "A2"

        # Mappings sheet should have frozen pane at A2
        mappings = wb["Bank Mappings"]
        assert mappings.freeze_panes == "A2"


class TestTransactionExportRowDTO:
    """Test suite for TransactionExportRowDTO."""

    def test_to_row_returns_correct_values(self):
        """Test that to_row() returns values in correct order."""
        dto = TransactionExportRowDTO(
            id="txn-1",
            date="2024-12-15",
            description="Test Transaction",
            counterparty="Test Merchant",
            amount=99.99,
            currency="EUR",
            debit_account="4200 - Groceries",
            credit_account="1000 - Checking",
            status="posted",
            source="bank_import",
            source_iban="DE89370400440532013000",
            counterparty_iban="DE11111111111111111111",
            is_internal_transfer=False,
            original_purpose="Original Purpose Text",
            bank_reference="BANK-REF-123",
            ai_suggested_account="Groceries",
            ai_confidence=0.85,
            ai_accepted=True,
        )

        row = dto.to_row()

        assert row[0] == "2024-12-15"  # date
        assert row[1] == "Test Transaction"  # description
        assert row[2] == "Test Merchant"  # counterparty
        assert row[3] == "DE11111111111111111111"  # counterparty_iban
        assert row[4] == 99.99  # amount
        assert row[5] == "EUR"  # currency
        assert row[11] == "No"  # is_internal_transfer
        assert row[14] == "Groceries"  # ai_suggested_account
        assert row[15] == "85%"  # ai_confidence formatted

    def test_to_row_handles_none_ai_values(self):
        """Test that to_row() handles None AI values gracefully."""
        dto = TransactionExportRowDTO(
            id="txn-1",
            date="2024-12-15",
            description="Test",
            counterparty="",
            amount=10.0,
            currency="EUR",
            debit_account="4200 - Test",
            credit_account="1000 - Test",
            status="posted",
            source="manual",
            source_iban="",
            counterparty_iban="",
            is_internal_transfer=False,
            original_purpose="",
            bank_reference="",
            ai_suggested_account="",
            ai_confidence=None,
            ai_accepted=None,
        )

        row = dto.to_row()

        assert row[14] == ""  # ai_suggested_account
        assert row[15] == ""  # ai_confidence (None formatted as empty)
        assert row[16] == ""  # ai_accepted (None formatted as empty)

    def test_column_headers_returns_expected_count(self):
        """Test that column_headers() returns the expected number of headers."""
        headers = TransactionExportRowDTO.column_headers()

        assert len(headers) == 17
        assert headers[0] == "Date"
        assert headers[-1] == "AI Accepted"


class TestMappingExportRowDTO:
    """Test suite for MappingExportRowDTO."""

    def test_to_row_formats_account_correctly(self):
        """Test that to_row() formats the linked account as 'number - name'."""
        dto = MappingExportRowDTO(
            iban="DE89370400440532013000",
            bank_account_name="Personal Checking",
            accounting_account_name="DKB Checking",
            accounting_account_number="1000",
            created_at="2024-01-01T00:00:00+00:00",
        )

        row = dto.to_row()

        assert row[0] == "DE89370400440532013000"
        assert row[1] == "Personal Checking"
        assert row[2] == "1000 - DKB Checking"  # Formatted as "number - name"
        assert row[3] == "2024-01-01T00:00:00+00:00"

    def test_column_headers_returns_expected_headers(self):
        """Test that column_headers() returns correct headers."""
        headers = MappingExportRowDTO.column_headers()

        assert headers == [
            "IBAN",
            "Bank Account Name",
            "Linked Accounting Account",
            "Created",
        ]
